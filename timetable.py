import logging
import telegram
import sqlite3
import datetime
import hashlib
import shutil
import os

from openpyxl import load_workbook
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)

logger = logging.getLogger(__name__)


# classes_list return list of classes prepared for making keyboard
def classes_list():
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    cursor_execution_result = cur.execute('select class_, letter from classes_ order by class_;').fetchall()
    keyboard = [[]]

    cursor_execution_result_size = len(cursor_execution_result)

    i = 0
    j = 0
    while i < cursor_execution_result_size:
        keyboard.append([])
        while j < cursor_execution_result_size and cursor_execution_result[j][0] == cursor_execution_result[i][0]:
            keyboard[i + 1].append(str(cursor_execution_result[j][0])
                                   + cursor_execution_result[j][1])
            j += 1
        i += 1

    con.close()
    return keyboard


# weekday_list return list of weekdays prepared for making keyboard
def weekday_list():
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    cursor_execution_result = cur.execute('select weekday from Weekdays order by id;').fetchall()
    keyboard = [[]]
    i = 0
    j = 0
    keyboard.append([])

    for row in cursor_execution_result:
        keyboard[i + 1].append(str(*cursor_execution_result[2 * i + j]))
        j += 1
        if j == 2:
            j = 0
            i += 1
            keyboard.append([])

    con.close()
    return keyboard


# teachers_list return list of teachers
def teachers_list():
    # take teachers list from database and make the list
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    cursor_execution_result = cur.execute('select SNF from Teachers order by SNF;').fetchall()
    teach_keyboard = []

    for row in cursor_execution_result:
        teach_keyboard.append(row)

    con.close()
    return teach_keyboard


def is_weekday(update, context):
    a = weekday_list()
    for i in range(0, len(a)):
        for j in range(0, len(a[i])):
            if update.message.text == a[i][j]:
                return True
    return False


def is_student(update, context):
    a = classes_list()
    for i in range(0, len(a)):
        for j in range(0, len(a[i])):
            if update.message.text == a[i][j]:
                return True
    return False


def is_teacher(update, context):
    a = teachers_list()
    for i in range(0, len(a)):
        if update.message.text == a[i][0]:
            return True
    return False


def class_selection_menu(update, context):
    markup = telegram.ReplyKeyboardMarkup([['Я учитель']] + classes_list(), one_time_keyboard=True)
    update.message.reply_text('Здравствуй!\nВыбери свой класс из списка ниже.', reply_markup=markup)


def teacher_selection_menu(update, context):
    teach_markup = telegram.ReplyKeyboardMarkup(teachers_list(), one_time_keyboard=True)
    update.message.reply_text('Выберите свое имя в списке ниже.\nНе учитель? Нажмите /start чтобы начать сначала.',
                              reply_markup=teach_markup)


def weekday_selection_menu(update, context):
    markup = telegram.ReplyKeyboardMarkup([['Сегодня'], ['Завтра']] + weekday_list() +
                                          [['Расписание звонков сегодня']] + [['Расписание звонков завтра']] +
                                          [['Свободные кабинеты']], one_time_keyboard=False)
    update.message.reply_text('Выбрать класс заново? /start', reply_markup=markup)


def new_teacher(update, context):
    teacher = context.args
    try:
        SNF = teacher[0] + ' ' + teacher[1] + ' ' + teacher[2]
        surname = teacher[0] + ' ' + teacher[1][0] + '.' + teacher[2][0] + '.'

        markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, добавить учителя']])
        update.message.reply_text('Вы дейстивтельно хотите добавить нового учителя?\n'
                                  f'_{SNF}_', reply_markup=markup, parse_mode='Markdown')
        teacher_to_add = [SNF, surname]
        with open(r'new.txt', 'w') as file:
            for line in teacher_to_add:
                file.write(line + '\n')
        file.close()
    except (IndexError):
        update.message.reply_text('Используйте /newt <Фамилия Имя Отчество>')


def add_new_teacher_to_base(update, context):
    teacher = []
    with open(r'new.txt', 'r') as file:
        for line in file:
            teacher.append(line)
    file.close()
    SNF = teacher[0]
    SNF.rstrip()
    surname = teacher[1]
    surname.rstrip()
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    cur.execute(f'INSERT INTO Teachers(SNF, surname_for_table) VALUES("{SNF}", "{surname}")')
    con.commit()
    con.close()
    fill_template_with_teachers(update, context)
    weekday_selection_menu(update, context)


def new_class(update, context):
    try:
        class_ = context.args[0]
        markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, добавить класс']])
        update.message.reply_text('Вы дейстивтельно хотите добавить новый класс?\n'
                                  f'_{class_}_', reply_markup=markup, parse_mode='Markdown')
        f = open(r'new.txt', 'w')
        f.write(class_)
        f.close()
    except (IndexError):
        update.message.reply_text('Используйте /newc <Класс>')


def add_new_class_to_base(update, context):
    f = open(r'new.txt', 'r')
    class_ = f.read()
    f.close()
    class_.rstrip()
    if len(class_) == 2:
        digit = class_[0]
        letter = class_[1]
    else:
        digit = class_[0] + class_[1]
        digit = int(digit)
        letter = class_[2]
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    cur.execute(f'INSERT INTO "main".Classes_(class_, letter) VALUES ({digit}, "{letter}")')
    con.commit()
    con.close()
    weekday_selection_menu(update, context)


def new_lesson(update, context):
    try:
        less = context.args
        lesson = ''
        for i in range(0, len(less)):
            lesson += less[i] + ' '
        lesson.rstrip()
        markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, добавить урок']])
        update.message.reply_text('Вы дейстивтельно хотите добавить новый урок?\n'
                                  f'_{lesson}_', reply_markup=markup, parse_mode='Markdown')
        f = open(r'new.txt', 'w')
        f.write(lesson)
        f.close()
    except (IndexError):
        update.message.reply_text('Используйте /newc <Урок>')


def add_new_lesson_to_base(update, context):
    f = open(r'new.txt', 'r')
    lesson = f.read()
    f.close()
    lesson.rstrip()
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    cur.execute(f'INSERT INTO Lessons(lesson) VALUES ("{lesson}");')
    con.commit()
    con.close()
    update.message.reply_text('Успешно')
    weekday_selection_menu(update, context)


# return id of class
def class_to_id(class_):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    if len(class_) == 3:
        letter = class_[2]
        digits = class_[0] + class_[1]
    else:
        letter = class_[1]
        digits = class_[0]
    uclass = cur.execute(
        f'select id from Classes_ where class_ == {int(digits)} and letter == "{letter}";').fetchone()
    con.close()
    return uclass[0]


def teacher_to_id(teacher):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    uteacher = cur.execute(f'select id from Teachers where SNF == "{str(teacher)}"').fetchone()
    con.close()
    return uteacher[0]


# connect user(chat_id) and teacher in table "Users"
def teachers(update, context):
    user = str(update.message.chat_id)
    cur_teacher = teacher_to_id(update.message.text)

    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    isreg = cur.execute(f'select class_or_teacher from Users where chat_id == {user};').fetchone()
    if (isreg == None):
        cur.execute(f'INSERT INTO main.Users(chat_id, class_or_teacher, is_teacher) VALUES ({user}, {cur_teacher}, 1);')
    else:
        cur.execute(f'UPDATE Users SET class_or_teacher = {cur_teacher}, is_teacher = 1 WHERE chat_id == {user}')

    con.commit()
    con.close()
    update.message.reply_text('Успешно!')
    weekday_selection_menu(update, context)


# connect user(chat_id) and class in table "Users"
def students(update, context):
    user = str(update.message.chat_id)
    cur_class = str(class_to_id(update.message.text))

    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    isreg = cur.execute(f'select class_or_teacher from Users where chat_id == {user};').fetchone()
    if (isreg == None):
        cur.execute(f'INSERT INTO main.Users(chat_id, class_or_teacher, is_teacher) VALUES ({user}, {cur_class}, 0);')
    else:
        cur.execute(f'UPDATE Users SET class_or_teacher = {cur_class}, is_teacher = 0 WHERE chat_id == {user}')

    con.commit()
    con.close()
    update.message.reply_text('Успешно!')
    weekday_selection_menu(update, context)


def preprinting(update, context, day):
    if day == 8:
        day = 1
    if day == 7:
        update.message.reply_text("Урааа, выходной)")
    else:
        con = sqlite3.connect('BOTSBASE.db')
        cur = con.cursor()
        # get id of weekday
        if day == 0:
            weekday_id = cur.execute(
                f'select id from Weekdays where weekday == "{update.message.text}";').fetchone()[0]
        else:
            weekday_id = day
        str_weekday = cur.execute(f'select weekday from "main".Weekdays where id == {weekday_id}').fetchone()[0]
        user = update.message.chat_id
        class_or_teacher = cur.execute(
            f'select class_or_teacher, is_teacher from Users WHERE chat_id == {user}').fetchone()
        con.close()
        if class_or_teacher == None:
            class_selection_menu(update, context)
        elif class_or_teacher[1] == 0:
            printing_for_students(update, context, weekday_id, class_or_teacher[0], str_weekday)
        else:
            printing_for_teachers(update, context, weekday_id, class_or_teacher[0], str_weekday)


def printing_for_teachers(update, context, weekday_id, teacher, str_weekday):
    con = sqlite3.connect("BOTSBASE.db")
    cur = con.cursor()

    timetable = cur.execute(
        f'select class_, cab, lesson, lesson_number from main_timetable where teacher == {teacher} and weekday == {weekday_id} order by lesson_number;'
    ).fetchall()
    str_teacher = cur.execute(f'select surname_for_table from "main".Teachers where id == {teacher}').fetchone()[0]

    keyboard = f'*{str_weekday} - {str_teacher}*\n\n'
    i = 0
    if len(timetable) == 0:
        keyboard += "Нет уроков"
    while i < len(timetable):
        class_ = cur.execute(f'select class_, letter from Classes_ where id == {timetable[i][0]};').fetchone()
        class_ = str(class_[0]) + class_[1]
        cab = str(timetable[i][1])
        lesson = cur.execute(f'select lesson from Lessons where id = {timetable[i][2]};').fetchone()
        lesson_number = str(timetable[i][3])
        keyboard += f'{lesson_number}) {class_} - {lesson[0]} ({cab})\n'
        i += 1

    con.close()
    update.message.reply_text(keyboard, parse_mode='Markdown')


def printing_for_students(update, context, weekday_id, class_, str_weekday):
    con = sqlite3.connect("BOTSBASE.db")
    cur = con.cursor()

    # get timetable for this weekday and class
    timetable = cur.execute(
        f'select lesson, cab, lesson_number, priority from main_timetable where weekday == {weekday_id} and class_ == {class_} order by lesson_number, priority;'
    ).fetchall()
    str_class = cur.execute(f'select class_, letter from Classes_ where id = {class_}').fetchone()
    str_class = str(str_class[0]) + str_class[1]

    keyboard = f'*{str_weekday} - {str_class}*\n\n'
    i = 0
    les_num = 1
    while i < len(timetable):
        keyboard += f'{timetable[i][2]}) '
        lesson_name = cur.execute(
            f'select lesson from Lessons where id=={timetable[i][0]};').fetchone()
        if len(timetable) - i > 1:
            if timetable[i][2] == timetable[i + 1][2]:
                if timetable[i][0] != timetable[i + 1][0]:
                    lesson_name2 = cur.execute(
                        f'select lesson from Lessons where id=={timetable[i + 1][0]}').fetchone()
                    keyboard += f"{lesson_name[0]} ({timetable[i][1]}) / {lesson_name2[0]} ({timetable[i + 1][1]})\n"
                else:
                    keyboard += f"{lesson_name[0]} ({timetable[i][1]}) / ({timetable[i + 1][1]})\n"
                i += 2
                les_num += 1
                continue
        if timetable[i][1] == 0:
            keyboard += f"{lesson_name[0]}\n"
        else:
            keyboard += f"{lesson_name[0]} ({timetable[i][1]})\n"
        i += 1
        les_num += 1

    con.close()
    update.message.reply_text(keyboard, parse_mode='Markdown')


def is_the_password_correct(the_password):  # может использовать много памяти
    return hashlib.pbkdf2_hmac('sha256', the_password.encode('UTF-8'), b'PTS', 50) == \
           b'\x89@\xf5Hd\x1a9\xd7\xb1\x07\x03\x81\x07b\xe1\x80\xd4h`/\xde\x16\xd6\x95\x9fUh\x05\xd7\x991w'
    # distanceBetweenSUNandEARTHequals0MB


def add_to_admins(update, context):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    user = update.message.chat_id
    cur.execute(f'INSERT INTO main.Admins(chat_id) VALUES ({user});')
    update.message.reply_text('Теперь вы можете вносить изменения в расписание набрав команду /edit')
    con.commit()
    con.close()


def is_admin(update, context):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    user = update.message.chat_id
    is_adm = cur.execute(f'select chat_id from Admins WHERE chat_id == {user}').fetchone()
    con.close()
    if is_adm == None:
        return False
    else:
        return True


def from_table_to_base(update, context):
    wb = load_workbook('Tables/present.xlsx')
    sheet = wb.active
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    teachers = cur.execute(
        f'select surname_for_table from Teachers order by surname_for_table;'
    ).fetchall()

    cur.execute(f'DELETE FROM main_timetable')

    for col in range(3, 3 * len(teachers) + 3, 3):
        teacher_cell = str(sheet.cell(row=1, column=col).value)
        teacher_id = cur.execute(f'select id from Teachers where surname_for_table == "{teacher_cell}"').fetchone()

        weekday = 1
        for week in range(2, 48, 9):
            for num in range(0, 9):
                lesson_cell = sheet.cell(column=col, row=week + num)
                lesson = lesson_cell.value
                if lesson != None:
                    prior = 0
                    id1 = lesson.find('1')
                    id2 = lesson.find('2')
                    if id1 != -1:
                        lesson = lesson[0:id1 - 1]
                        lesson.rstrip()
                        prior = 1
                    elif id2 != -1:
                        lesson = lesson[0:id2 - 1]
                        lesson.rstrip()
                        prior = 2

                    lesson = cur.execute(f'select id from Lessons WHERE lesson == "{lesson}"').fetchone()

                    cab_cell = sheet.cell(column=col + 2, row=week + num)
                    cab = cab_cell.value

                    # filling 'Cab' table
                    cabinet = cur.execute(f'select id from Cabs WHERE cab == {cab};').fetchone()
                    if cabinet == None:
                        cur.execute(f'INSERT INTO Cabs(cab) VALUES ({cab});')

                    class_cell = sheet.cell(column=col + 1, row=week + num)
                    clas = class_cell.value
                    if len(clas) == 3:
                        digit = clas[0] + clas[1]
                        letter = clas[2]
                    else:
                        digit = clas[0]
                        letter = clas[1]
                    clas = cur.execute(
                        f'select id from Classes_ WHERE class_ == {int(digit)} AND letter == "{letter}"').fetchone()

                    lesson_num = sheet.cell(column=2, row=week + num).value
                    # print(teacher_id[0], clas[0], cab, lesson[0], weekday, lesson_num) logs
                    cur.execute(
                        f'INSERT INTO main_timetable (teacher, class_, cab, lesson, weekday, lesson_number, priority) '
                        f'VALUES ({teacher_id[0]}, {clas[0]}, {cab}, {lesson[0]}, {weekday}, {lesson_num}, {prior});'
                    )

            weekday += 1
    con.commit()
    con.close()
    update.message.reply_text('Успешно')


# previous name is from_base_to_table
def fill_template_with_teachers(update, context):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    wb = load_workbook('Tables/template.xlsx')
    # will use copy of template
    sheet = wb.active

    teachers = cur.execute(
        f'select surname_for_table from Teachers order by surname_for_table'
    ).fetchall()

    # fill row with teachers
    tcol = 3
    for col in range(3, len(teachers) + 3):
        value = teachers[col - 3][0]
        sheet.merge_cells(start_row=1, end_row=1, start_column=tcol, end_column=tcol + 2)
        cell = sheet.cell(row=1, column=tcol)
        cell.value = value
        tcol += 3

    wb.save('Tables/template.xlsx')


def download_photo(update, context):
    if is_admin(update, context):
        photo_file = update.message.photo[-1].get_file()
        photo_file.download('user_photo.png')
        confirm_to_send_photo(update, context)


def confirm_to_send_photo(update, context):
    markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, отправить фото']])
    update.message.reply_text('Вы дейстивтельно хотите отправить это фото всем пользователям?', reply_markup=markup)


def send_to_all_users(update, context, arg):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    users = cur.execute(f'select chat_id from "main".Users').fetchall()
    if arg == 0:
        for i in range(0, len(users)):
            context.bot.send_photo(chat_id=users[i][0], photo=open('user_photo.png', 'rb'))
    if arg == 1:
        f = open('alert.txt', 'r')
        text = f.read()
        f.close()
        for i in range(0, len(users)):
            context.bot.send_message(chat_id=users[i][0], text=text)
    weekday_selection_menu(update, context)


def confirm_to_send_text(update, context, text):
    if text[1] == ' ':
        text = text[2:]
    else:
        text = text[1:]
    markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, отправить текст']])
    update.message.reply_text('Вы дейстивтельно хотите отправить текст ниже всем пользователям?\n\n'
                              f'_{text}_', reply_markup=markup, parse_mode="Markdown")
    f = open('alert.txt', 'w')
    f.write(text)
    f.close()


def download_document(update, context):
    if is_admin(update, context):
        doc_file = update.message.document.get_file()
        doc_file.download('user_doc.xlsx')
        markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, изменить постоянное расписание']])
        update.message.reply_text('Вы дейстивтельно хотите изменить нынешнее расписани на то что вы прислали?\n\n',
                                  reply_markup=markup, parse_mode="Markdown")


def change_main_timetable(update, context):
    date = datetime.datetime.today().date()
    date = str(date).replace('-', '')
    time = datetime.datetime.today().time().replace(second=0, microsecond=0)
    time = str(time).replace(':', '')
    name_of_reserve_copy = 'Копия_от_' + str(date) + '_' + str(time) + '.xlsx'

    shutil.copyfile('Tables/present.xlsx', f'Tables/just in case/{name_of_reserve_copy}')

    shutil.copyfile('Tables/present.xlsx', 'Tables/past.xlsx')
    shutil.copyfile('user_doc.xlsx', 'Tables/present.xlsx')
    from_table_to_base(update, context)


def send_table_to_admin(update, context, arg):
    if arg == 'present':
        table = open('Tables/present.xlsx', 'rb')
    if arg == 'past':
        table = open('Tables/past.xlsx', 'rb')
    if arg == 'all':
        update.message.reply_text('расшифровка чисел в названии - ГГГГММДД-ЧЧММ')
        directory = 'Tables/just in case'
        files = os.listdir(directory)
        for i in files:
            table = open(f'Tables/just in case/{i}', 'rb')
            context.bot.send_document(chat_id=update.message.chat_id, document=table)
    if arg != 'all':
        context.bot.send_document(chat_id=update.message.chat_id, document=table)


def new_timetable(update, context):
    update.message.reply_text('Заполните данную таблицу и отправьте ее сюда в любое время\n'
                              'Перед началом работы *обязательно* ознакомьтесь с инструкцией', parse_mode='Markdown')
    table_file = open(r'Tables/template.xlsx', 'rb')
    context.bot.send_document(update.message.chat_id, table_file)

    instruction_file = open(r'instruction for filling timetable.txt', 'rb')
    context.bot.send_document(update.message.chat_id, instruction_file)

    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    lessons = cur.execute('select lesson from Lessons').fetchall()
    con.close()
    text = ''
    for i in range(0, len(lessons)):
        text += lessons[i][0] + '\n'
    f = open(r'lessons_list.txt', 'w')
    f.write(text)
    f.close()

    list_file = open(r'lessons_list.txt', 'rb')
    context.bot.send_document(update.message.chat_id, list_file)


def pre_delete_teacher(update, context):
    text = 'Для удаления учителя введите */delt <Фамилия Имя Отчество>*\n' \
           'Пример: _/delt Иванов Иван Иванович_\n\n' \
           'Список учителей:\n'
    teachers = teachers_list()
    for i in range(0, len(teachers)):
        text += teachers[i][0] + '\n'
    update.message.reply_text(text, parse_mode='Markdown')


def confirm_to_delete_teacher(update, context):
    teach = context.args
    teacher = ''
    for i in range(0, len(teach)):
        teacher += teach[i] + ' '
    teacher.rstrip()
    markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, удалить учителя']])
    update.message.reply_text(f'Вы дейстивтельно хотите удалить этого учителя?\n'
                              f'_{teacher}_', reply_markup=markup, parse_mode="Markdown")
    f = open(r'new.txt', 'w')
    f.write(teacher)
    f.close()


def delete_teacher(update, context):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    f = open(r'new.txt', 'r')
    teacher = f.read()
    f.close()
    try:
        a = cur.execute(f'select id from Teachers WHERE SNF == "{teacher}"').fetchone()
        a = int(a[0])
        cur.execute(f'DELETE FROM Teachers WHERE id == {a};')
        con.commit()
        con.close()
        # fill_template_with_teachers(update, context)
        weekday_selection_menu(update, context)
    except (sqlite3.OperationalError, TypeError):
        update.message.reply_text("Нет такого учителя")
        pre_delete_teacher(update, context)


def pre_delete_class(update, context):
    text = 'Для удаления класса введите */delc <Класс>*\n' \
           'Пример: _/delc 5А_\n\n' \
           'Список классов:\n'
    classes = classes_list()
    for i in range(0, len(classes)):
        for j in range(0, len(classes[i])):
            text += classes[i][j] + '\n'
    update.message.reply_text(text, parse_mode='Markdown')


def confirm_to_delete_class(update, context):
    class_ = context.args
    class_ = class_[0]
    markup = telegram.ReplyKeyboardMarkup([['Нет']] + [['Да, удалить класс']])
    update.message.reply_text(f'Вы дейстивтельно хотите удалить этот класс?\n'
                              f'_{class_}_', reply_markup=markup, parse_mode="Markdown")
    f = open(r'new.txt', 'w')
    f.write(class_)
    f.close()


def delete_class(update, context):
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    f = open(r'new.txt', 'r')
    class_ = f.read()
    f.close()
    digit = -1
    letter = -1
    if len(class_) == 3:
        digit = int(class_[0] + class_[1])
        letter = class_[2]
    if len(class_) == 2:
        digit = int(class_[0])
        letter = class_[1]
    try:
        a = cur.execute(f'select id from Classes_ WHERE class_ == {digit} AND letter == "{letter}"').fetchone()
        a = int(a[0])
        cur.execute(f'DELETE FROM Classes_ WHERE id == {a};')
        con.commit()
        con.close()
        # fill_template_with_teachers(update, context)
        weekday_selection_menu(update, context)
    except (sqlite3.OperationalError, TypeError):
        update.message.reply_text("Нет такого класса")
        pre_delete_class(update, context)


def editing(update, context):
    if is_admin(update, context):
        markup = telegram.ReplyKeyboardMarkup(
            [['Изменить постоянное расписание']] +
            [['Скачать таблицу с расписанием']] +
            [['Скачать прошлое расписание']] +
            [['Получить всю базу расписаний']] +
            [['Удалить учителя']] +
            [['Удалить класс']] +
            [['Выйти из этого меню']])
        update.message.reply_text('МЕНЮ редактирования расписания', reply_markup=markup)

        update.message.reply_text(
            'Нажав кнопку *Изменить постоянное расписание*, вы получите таблицу, '
            'которую нужно будет заполнить и отправить сюда находясь в любом меню '
            '(Просьба заполнять аккуратно, соблюдая приложенную инструкцию)\n'
            'Если что-то испортили, можно скачать таблицу с расписанием установленным здесь прежде, '
            'нажав кнопку *Скачать прошлое расписание*\n'
            'Если хотите получить таблицу с нынешним расписанием, нажмите *Скачать таблицу с расписанием*\n'
            'А если уж все совсем плохо, можно скачать все расписания которые когда-либо были тут. '
            'Для этого нажмите *Получить всю базу расписаний*\n\n'

            'Вы можете отправить сообщение всем пользователям:\n'
            '1) Загрузите фотографию в любом меню бота и подтвердите отправку\n'
            '2) Отправьте любой текст с *!* в начале\n'
            'Пример:\n'
            ' _!Привет всем пользоватлеям бота_\n\n'

            'Вы можете добавить нового учителя написав */newt <Фамилия Имя Отчество>*\n'
            'Пример:\n'
            '_/newt Иванов Иван Иванович_\n\n'

            'Вы можете добавить новый класс написав */newc <Класс>*\n'
            'Пример:\n'
            '_/newc 5А_\n'
            'При добавлении класса просьба *не ставить пробел* между числом и буквой\n\n'
            'Вы можете добавить новый урок написав */newl <Урок>*\n'
            'Пример: \n'
            '_/newl Русский язык_\n\n'
            'Удалить учителя/класс можно нажав соответствующую кнопку\n\n'

            'Если возникли ошибки или появились предложения, пишите ему @nefizik'

            , parse_mode='Markdown')


def ring_schedule(update, context, arg):
    if arg == 8:
        arg = 1
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()

    text = '*Расписание звонков - '
    if arg == 7:
        text += 'Воскресенье*'
    else:
        text += cur.execute(f'select weekday from Weekdays where id == {int(arg)}').fetchone()[0] + '*'
    text += '\n\n'
    text += cur.execute(f'select schedule from Rings_schedule WHERE weekday == {int(arg)}').fetchone()[0]
    update.message.reply_text(text, parse_mode="Markdown")


def empty_classrooms(update, context, arg):
    if arg == 7:
        update.message.reply_text('Сегодня все кабинеты свободны)')
        return
    con = sqlite3.connect('BOTSBASE.db')
    cur = con.cursor()
    weekday = cur.execute(f'select weekday from Weekdays WHERE id == {arg}').fetchone()[0]
    text = '*Свободные кабинеты - ' + weekday + '*\n\n'
    for k in range(1, 10):
        text += str(k) + ') '
        all_classrooms = cur.execute(f'select cab from Cabs').fetchall()
        busy_classrooms = cur.execute(
            f'select cab from main_timetable WHERE weekday == {arg} AND lesson_number == {k};').fetchall()
        for i in range(0, len(busy_classrooms)):
            for j in range(0, len(all_classrooms)):
                if busy_classrooms[i] == all_classrooms[j]:
                    all_classrooms.pop(j)
                    break
        all_classrooms.sort()
        for i in range(0, len(all_classrooms)):
            if all_classrooms[i][0] == 0:
                continue
            text += str(all_classrooms[i][0]) + ", "
        text = text[:-2]
        text += '\n'
    con.close()
    update.message.reply_text(text, parse_mode='Markdown')


def distributor(update, context):
    ini = update.message.text
    if ini == 'Я учитель':
        teacher_selection_menu(update, context)
    elif is_student(update, context):
        students(update, context)
    elif is_teacher(update, context):
        teachers(update, context)
    elif is_weekday(update, context):
        day = 0
        preprinting(update, context, day)
    elif ini == 'Сегодня':
        day = datetime.datetime.today().isoweekday()
        preprinting(update, context, day)
    elif ini == 'Завтра':
        day = datetime.datetime.today().isoweekday() + 1
        preprinting(update, context, day)
    elif ini == 'Расписание звонков сегодня':
        day = datetime.datetime.today().isoweekday()
        ring_schedule(update, context, day)
    elif ini == 'Расписание звонков завтра':
        day = datetime.datetime.today().isoweekday() + 1
        ring_schedule(update, context, day)
    elif ini == 'Свободные кабинеты':
        day = datetime.datetime.today().isoweekday()
        empty_classrooms(update, context, day)

    # admin`s commands
    elif is_the_password_correct(update.message.text):
        add_to_admins(update, context)
    if is_admin(update, context):
        if ini == 'Изменить постоянное расписание':
            new_timetable(update, context)
        elif ini == 'Скачать таблицу с расписанием':
            send_table_to_admin(update, context, "present")
        elif ini == 'Скачать прошлое расписание':
            send_table_to_admin(update, context, 'past')
        elif ini == 'Получить всю базу расписаний':
            send_table_to_admin(update, context, 'all')
        elif ini == 'Удалить учителя':
            pre_delete_teacher(update, context)
        elif ini == 'Удалить класс':
            pre_delete_class(update, context)
        elif ini == 'Нет':
            update.message.reply_text("Изменения не внесены")
            weekday_selection_menu(update, context)
        elif ini == 'Да, отправить фото':
            send_to_all_users(update, context, 0)
        elif ini == 'Да, отправить текст':
            send_to_all_users(update, context, 1)
        elif ini == 'Да, добавить учителя':
            add_new_teacher_to_base(update, context)
        elif ini == 'Да, добавить класс':
            add_new_class_to_base(update, context)
        elif ini == 'Да, добавить урок':
            add_new_lesson_to_base(update, context)
        elif ini == 'Да, изменить постоянное расписание':
            change_main_timetable(update, context)
        elif ini == 'Да, удалить учителя':
            delete_teacher(update, context)
        elif ini == 'Да, удалить класс':
            delete_class(update, context)
        elif ini == 'Выйти из этого меню':
            weekday_selection_menu(update, context)
        elif ini[0] == '!':
            confirm_to_send_text(update, context, ini)


def main():
    updater = Updater(token='884658566:AAG5l3pY4CacvQamPZBAJhF25NjUyQHnp4k', use_context=True)

    # get the dispatcher to register handlers (обработчики)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler('start', class_selection_menu))
    dp.add_handler(CommandHandler('edit', editing))
    dp.add_handler(CommandHandler('newt', new_teacher, pass_chat_data=True))
    dp.add_handler(CommandHandler('newc', new_class, pass_chat_data=True))
    dp.add_handler(CommandHandler('newl', new_lesson, pass_chat_data=True))
    dp.add_handler(CommandHandler('delt', confirm_to_delete_teacher, pass_chat_data=True))
    dp.add_handler(CommandHandler('delc', confirm_to_delete_class, pass_chat_data=True))

    dp.add_handler(MessageHandler(filters=Filters.text, callback=distributor))
    dp.add_handler(MessageHandler(filters=Filters.photo, callback=download_photo))
    dp.add_handler(MessageHandler(filters=Filters.document, callback=download_document))
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()
